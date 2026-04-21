import math
import os
import subprocess
import sys
import tempfile
import threading
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, List, Optional, Sequence, Set, Tuple

import ezdxf
import openpyxl
from ezdxf import recover
from ezdxf.math import Vec2, Vec3, bulge_to_arc
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ProgressCallback = Callable[[int, int], None]
ResultCallback = Callable[["FileMeasureResult", int, int, int, int], None]
StateCallback = Callable[[str, int, int], None]
DiscoveryCallback = Callable[[int, int], None]
FolderCallback = Callable[[str, int, int], None]


@dataclass(frozen=True)
class ScanOptions:
    source_path: Path
    target_layers: Tuple[str, ...]
    unit_to_mm_factor: float = 1.0
    recursive: bool = True
    workers: int = 4
    include_missing_layers: bool = True
    include_read_errors: bool = True
    spline_segments: int = 96
    ellipse_segments: int = 96


@dataclass(frozen=True)
class FileMeasureResult:
    filename: str
    full_path: str
    folder_rel: str
    measures_mm: Dict[str, Optional[float]]
    status: str


@dataclass(frozen=True)
class FileTask:
    path: Path
    folder_rel: str
    folder_index: int
    folder_count: int


class ScanController:
    def __init__(self) -> None:
        self._pause_event = threading.Event()
        self._stop_event = threading.Event()

    def pause(self) -> None:
        self._pause_event.set()

    def resume(self) -> None:
        self._pause_event.clear()

    def stop(self) -> None:
        self._stop_event.set()
        self._pause_event.clear()

    def is_paused(self) -> bool:
        return self._pause_event.is_set()

    def is_stopped(self) -> bool:
        return self._stop_event.is_set()

    def wait_if_paused(self, poll_interval: float = 0.05) -> bool:
        while self.is_paused():
            if self.is_stopped():
                return False
            threading.Event().wait(poll_interval)
        return not self.is_stopped()


def normalize_layers(values: Sequence[str]) -> Tuple[str, ...]:
    normalized: List[str] = []
    seen: Set[str] = set()

    for raw in values:
        value = str(raw).strip()
        if not value:
            continue
        key = value.casefold()
        if key in seen:
            continue
        seen.add(key)
        normalized.append(value)

    if not normalized:
        raise ValueError("No valid layer was provided.")

    return tuple(normalized)


def polyline_length(points: Sequence[Vec2 | Vec3]) -> float:
    total = 0.0
    for a, b in zip(points, points[1:]):
        total += (b - a).magnitude
    return float(total)


def _arc_length_from_bulge(start: Vec2, end: Vec2, bulge: float) -> float:
    if abs(bulge) < 1e-12:
        return float((end - start).magnitude)
    _, start_angle, end_angle, radius = bulge_to_arc(start, end, bulge)
    sweep = (end_angle - start_angle) % (2.0 * math.pi)
    return float(radius * sweep)


def _lwpolyline_length(entity) -> float:
    points = list(entity.get_points("xyb"))
    if len(points) < 2:
        return 0.0

    total = 0.0
    segment_count = len(points) if entity.closed else len(points) - 1
    for index in range(segment_count):
        x1, y1, bulge = points[index]
        x2, y2, _ = points[(index + 1) % len(points)]
        total += _arc_length_from_bulge(Vec2(x1, y1), Vec2(x2, y2), float(bulge))
    return float(total)


def _polyline_length(entity) -> float:
    vertices = list(entity.vertices)
    if len(vertices) < 2:
        return 0.0

    if entity.is_2d_polyline:
        total = 0.0
        segment_count = len(vertices) if entity.is_closed else len(vertices) - 1
        for index in range(segment_count):
            start = Vec2(vertices[index].dxf.location)
            end = Vec2(vertices[(index + 1) % len(vertices)].dxf.location)
            bulge = float(getattr(vertices[index].dxf, "bulge", 0.0) or 0.0)
            total += _arc_length_from_bulge(start, end, bulge)
        return float(total)

    points = [Vec3(vertex.dxf.location) for vertex in vertices]
    total = polyline_length(points)
    if entity.is_closed and len(points) > 1:
        total += (points[0] - points[-1]).magnitude
    return float(total)


def entity_length(entity, spline_segments: int, ellipse_segments: int) -> float:
    entity_type = entity.dxftype()

    if entity_type == "LINE":
        start = Vec3(entity.dxf.start)
        end = Vec3(entity.dxf.end)
        return float((end - start).magnitude)

    if entity_type == "LWPOLYLINE":
        return _lwpolyline_length(entity)

    if entity_type == "POLYLINE":
        return _polyline_length(entity)

    if entity_type == "ARC":
        radius = float(entity.dxf.radius)
        start_angle = float(entity.dxf.start_angle)
        end_angle = float(entity.dxf.end_angle)
        sweep_deg = (end_angle - start_angle) % 360.0
        return float(radius * math.radians(sweep_deg))

    if entity_type == "CIRCLE":
        radius = float(entity.dxf.radius)
        return float(2.0 * math.pi * radius)

    if entity_type == "SPLINE":
        try:
            points = [Vec3(point) for point in entity.construction_tool().approximate(segments=spline_segments)]
            return float(polyline_length(points))
        except Exception:
            return 0.0

    if entity_type == "ELLIPSE":
        try:
            points = [Vec3(point) for point in entity.construction_tool().approximate(segments=ellipse_segments)]
            return float(polyline_length(points))
        except Exception:
            return 0.0

    if entity_type == "INSERT":
        total = 0.0
        try:
            for virtual_entity in entity.virtual_entities():
                total += entity_length(virtual_entity, spline_segments, ellipse_segments)
        except Exception:
            return float(total)
        return float(total)

    return 0.0


def safe_read_dxf(path: Path) -> Tuple[Optional["ezdxf.document.Drawing"], Optional[str]]:
    try:
        return ezdxf.readfile(str(path)), None
    except Exception:
        try:
            doc, _ = recover.readfile(str(path))
            return doc, None
        except Exception as exc:
            return None, type(exc).__name__


def _folder_rel(base: Path, folder: Path) -> str:
    try:
        rel = str(folder.relative_to(base))
        return "." if rel == "." else rel.replace("\\", "/")
    except Exception:
        return str(folder).replace("\\", "/")


def collect_dxf_tasks(source_path: Path, recursive: bool) -> Tuple[List[FileTask], int]:
    if source_path.is_file():
        if source_path.suffix.lower() != ".dxf":
            raise ValueError(f"Selected file is not a DXF: {source_path}")
        return [FileTask(path=source_path, folder_rel=".", folder_index=1, folder_count=1)], 1

    if not source_path.is_dir():
        raise FileNotFoundError(f"Path not found: {source_path}")

    folders: List[Tuple[Path, List[Path]]] = []

    if recursive:
        for dirpath, _, filenames in os.walk(source_path):
            dxf_files = [
                Path(dirpath) / name
                for name in filenames
                if name.lower().endswith(".dxf")
            ]
            if dxf_files:
                dxf_files.sort(key=lambda item: item.name.lower())
                folders.append((Path(dirpath), dxf_files))
    else:
        dxf_files = [
            source_path / name
            for name in os.listdir(source_path)
            if (source_path / name).is_file() and name.lower().endswith(".dxf")
        ]
        dxf_files.sort(key=lambda item: item.name.lower())
        if dxf_files:
            folders.append((source_path, dxf_files))

    if not folders:
        raise FileNotFoundError("No DXF file was found.")

    tasks: List[FileTask] = []
    folder_count = len(folders)

    for idx, (folder_path, files) in enumerate(sorted(folders, key=lambda item: str(item[0]).lower()), start=1):
        rel = _folder_rel(source_path, folder_path)
        for path in files:
            tasks.append(FileTask(path=path, folder_rel=rel, folder_index=idx, folder_count=folder_count))

    return tasks, folder_count


def measure_layers_in_document(
    doc: "ezdxf.document.Drawing",
    target_layers: Sequence[str],
    unit_to_mm_factor: float,
    spline_segments: int,
    ellipse_segments: int,
) -> Tuple[Dict[str, Optional[float]], str]:
    layer_lookup = {layer.casefold(): layer for layer in target_layers}
    totals = {layer: 0.0 for layer in target_layers}
    seen_layers: Set[str] = set()

    for entity in doc.modelspace():
        layer_name = getattr(entity.dxf, "layer", None)
        if layer_name is None:
            continue
        requested_layer = layer_lookup.get(str(layer_name).casefold())
        if requested_layer is None:
            continue
        seen_layers.add(requested_layer)
        totals[requested_layer] += entity_length(entity, spline_segments, ellipse_segments)

    measures_mm: Dict[str, Optional[float]] = {}
    missing_layers: List[str] = []

    for layer in target_layers:
        if layer in seen_layers:
            measures_mm[layer] = round(totals[layer] * unit_to_mm_factor, 3)
        else:
            measures_mm[layer] = None
            missing_layers.append(layer)

    if len(missing_layers) == len(target_layers):
        status = "NO LAYER FOUND"
    elif missing_layers:
        status = "PARTIAL"
    else:
        status = "OK"

    return measures_mm, status


def process_one_file(
    dxf_path: str,
    folder_rel: str,
    target_layers: Tuple[str, ...],
    unit_to_mm_factor: float,
    spline_segments: int,
    ellipse_segments: int,
) -> FileMeasureResult:
    path = Path(dxf_path)
    doc, error = safe_read_dxf(path)

    if doc is None:
        return FileMeasureResult(
            filename=path.name,
            full_path=str(path.resolve(strict=False)),
            folder_rel=folder_rel,
            measures_mm={layer: None for layer in target_layers},
            status=f"ERROR:{error}",
        )

    try:
        measures_mm, status = measure_layers_in_document(
            doc=doc,
            target_layers=target_layers,
            unit_to_mm_factor=unit_to_mm_factor,
            spline_segments=spline_segments,
            ellipse_segments=ellipse_segments,
        )
        return FileMeasureResult(
            filename=path.name,
            full_path=str(path.resolve(strict=False)),
            folder_rel=folder_rel,
            measures_mm=measures_mm,
            status=status,
        )
    except Exception as exc:
        return FileMeasureResult(
            filename=path.name,
            full_path=str(path.resolve(strict=False)),
            folder_rel=folder_rel,
            measures_mm={layer: None for layer in target_layers},
            status=f"ERROR:{type(exc).__name__}",
        )


def should_keep_result(result: FileMeasureResult, options: ScanOptions) -> bool:
    if result.status.startswith("ERROR:") and not options.include_read_errors:
        return False
    if result.status == "NO LAYER FOUND" and not options.include_missing_layers:
        return False
    return True


def status_details(result: FileMeasureResult, layers: Sequence[str]) -> str:
    measured: List[str] = []
    missing: List[str] = []

    for layer in layers:
        value = result.measures_mm.get(layer)
        if value is None:
            missing.append(layer)
        else:
            measured.append(f"{layer}={value:.3f} mm")

    parts = [result.status]
    if measured:
        parts.append("found: " + ", ".join(measured))
    if missing and not result.status.startswith("ERROR:"):
        parts.append("missing: " + ", ".join(missing))
    return " | ".join(parts)


def _apply_header_style(worksheet) -> None:
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    font = Font(bold=True, color="1F1F1F")
    border = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )

    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def _style_body(worksheet, first_layer_col: int, last_layer_col: int, full_path_col: int) -> None:
    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx in range(first_layer_col, last_layer_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.000" mm"'
                cell.alignment = Alignment(horizontal="right")

        path_cell = worksheet.cell(row=row_idx, column=full_path_col)
        path_cell.alignment = Alignment(horizontal="left")


def _fit_columns(worksheet) -> None:
    for idx, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        header = str(worksheet.cell(row=1, column=idx).value or "")
        if header == "File":
            width = min(max(max_length + 2, 22), 42)
        elif header == "Full path":
            width = min(max(max_length + 2, 48), 90)
        else:
            width = min(max(max_length + 2, 12), 18)

        worksheet.column_dimensions[get_column_letter(idx)].width = width


def write_results(output_path: Path, layers: Sequence[str], results: Sequence[FileMeasureResult]) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Measurements"

    headers = ["File", *[f"Layer {layer}" for layer in layers], "Full path"]
    worksheet.append(headers)

    for result in results:
        row = [result.filename]
        for layer in layers:
            row.append(result.measures_mm.get(layer))
        row.append(result.full_path)
        worksheet.append(row)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = True

    _apply_header_style(worksheet)
    _style_body(
        worksheet,
        first_layer_col=2,
        last_layer_col=1 + len(layers),
        full_path_col=2 + len(layers),
    )
    _fit_columns(worksheet)

    workbook.save(output_path)


def create_temp_output_path(prefix: str = "dxf_layer_report_") -> Path:
    handle = tempfile.NamedTemporaryFile(prefix=prefix, suffix=".xlsx", delete=False)
    handle.close()
    return Path(handle.name)


def open_in_default_spreadsheet(path: Path) -> None:
    if sys.platform.startswith("win"):
        os.startfile(str(path))  # type: ignore[attr-defined]
        return
    if sys.platform == "darwin":
        subprocess.Popen(["open", str(path)])
        return
    subprocess.Popen(["xdg-open", str(path)])


def scan_dxf_files(
    options: ScanOptions,
    progress_callback: Optional[ProgressCallback] = None,
    result_callback: Optional[ResultCallback] = None,
    controller: Optional[ScanController] = None,
    state_callback: Optional[StateCallback] = None,
    discovery_callback: Optional[DiscoveryCallback] = None,
    folder_callback: Optional[FolderCallback] = None,
) -> Tuple[List[FileMeasureResult], int, bool]:
    tasks, folder_count = collect_dxf_tasks(options.source_path, options.recursive)
    total_files = len(tasks)

    if discovery_callback:
        discovery_callback(folder_count, total_files)
    if progress_callback:
        progress_callback(0, total_files)

    results: List[FileMeasureResult] = []
    done = 0
    stopped_early = False
    last_folder_rel = None

    def emit_folder(task: FileTask) -> None:
        nonlocal last_folder_rel
        if folder_callback and task.folder_rel != last_folder_rel:
            last_folder_rel = task.folder_rel
            folder_callback(task.folder_rel, task.folder_index, task.folder_count)

    if options.workers <= 1 or total_files <= 1:
        for task in tasks:
            if controller:
                if controller.is_stopped():
                    stopped_early = True
                    break
                if controller.is_paused():
                    if state_callback:
                        state_callback("paused", done, total_files)
                    if not controller.wait_if_paused():
                        stopped_early = True
                        break
                    if state_callback:
                        state_callback("running", done, total_files)

            emit_folder(task)

            result = process_one_file(
                dxf_path=str(task.path),
                folder_rel=task.folder_rel,
                target_layers=options.target_layers,
                unit_to_mm_factor=options.unit_to_mm_factor,
                spline_segments=options.spline_segments,
                ellipse_segments=options.ellipse_segments,
            )
            done += 1
            if should_keep_result(result, options):
                results.append(result)
            if result_callback:
                result_callback(result, done, total_files, task.folder_index, task.folder_count)
            if progress_callback:
                progress_callback(done, total_files)
    else:
        max_workers = max(1, min(int(options.workers), 32, total_files))
        inflight_limit = max_workers * 3
        next_index = 0
        pending: Dict[object, FileTask] = {}

        with ThreadPoolExecutor(max_workers=max_workers, thread_name_prefix="DXFWorker") as executor:
            while True:
                if controller and controller.is_stopped():
                    stopped_early = True
                    break

                if controller and controller.is_paused():
                    if state_callback:
                        state_callback("paused", done, total_files)
                    if not controller.wait_if_paused():
                        stopped_early = True
                        break
                    if state_callback:
                        state_callback("running", done, total_files)

                while next_index < total_files and len(pending) < inflight_limit:
                    task = tasks[next_index]
                    emit_folder(task)
                    future = executor.submit(
                        process_one_file,
                        str(task.path),
                        task.folder_rel,
                        options.target_layers,
                        options.unit_to_mm_factor,
                        options.spline_segments,
                        options.ellipse_segments,
                    )
                    pending[future] = task
                    next_index += 1

                if not pending:
                    if next_index >= total_files:
                        break
                    continue

                done_set, _ = wait(tuple(pending.keys()), timeout=0.08, return_when=FIRST_COMPLETED)
                if not done_set:
                    continue

                for future in done_set:
                    task = pending.pop(future)
                    result = future.result()
                    done += 1
                    if should_keep_result(result, options):
                        results.append(result)
                    if result_callback:
                        result_callback(result, done, total_files, task.folder_index, task.folder_count)
                    if progress_callback:
                        progress_callback(done, total_files)

            if stopped_early and pending:
                for future in pending:
                    future.cancel()

    results.sort(key=lambda item: (item.folder_rel.lower(), item.filename.lower(), item.full_path.lower()))
    return results, total_files, stopped_early
